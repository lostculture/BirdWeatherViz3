"""
Settings API Endpoints
Endpoints for application configuration and settings.

Version: 1.0.0
"""

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from typing import List, Optional, Dict
from pydantic import BaseModel
import csv
import io
import os
import re
import shutil
import tempfile
import time
from pathlib import Path

from app.api.deps import get_db_dependency, get_current_user
from app.config import settings as app_settings
from app.db.models.setting import Setting
from app.db.models.species import Species
from app.db.models.station import Station
from app.db.models.detection import Detection
from app.db.models.taxonomy_translation import TaxonomyTranslation
from app.services import taxonomy_translations as _tx
from datetime import datetime, timedelta
import logging

logger = logging.getLogger(__name__)

router = APIRouter()


class SettingResponse(BaseModel):
    """Setting response model."""
    key: str
    value: Optional[str]
    data_type: Optional[str]
    description: Optional[str]


class SettingUpdate(BaseModel):
    """Setting update model."""
    value: str
    data_type: Optional[str] = "str"
    description: Optional[str] = None


class TaxonomyUploadResponse(BaseModel):
    """Response model for taxonomy upload."""
    success: bool
    species_updated: int
    species_created: int
    families_added: int
    total_ebird_species: int
    languages_imported: List[str] = []
    translations_imported: int = 0
    message: str


# Any column in the uploaded taxonomy file whose header (case-insensitive) is
# in this set is a known taxonomic / housekeeping column and is NOT a language
# column. Everything else is treated as a per-language common-name column, with
# the header used verbatim as the language label. eBird's multilingual v2025+
# XLSX uses human-readable names like "Italian" or "Spanish, Mexico" as column
# headers — storing those verbatim makes the Configuration dropdown friendlier
# for non-technical users than ISO codes would be.
_CORE_COLUMNS = {
    # Identifiers
    "SPECIES_CODE", "SPECIES_CD", "SPECIESCODE",
    "SCI_NAME", "SCIENTIFIC_NAME", "SCINAME",
    "PRIMARY_COM_NAME", "COMMON_NAME", "COMNAME",
    # Taxonomy hierarchy
    "ORDER1", "ORDER_NAME", "ORDER",
    "FAMILY", "FAMILY_NAME",
    "SPECIES_GROUP",
    # eBird housekeeping columns that appear in various sheet variants
    "CATEGORY", "TAXON_ORDER",
    "REPORT_AS", "SPARSE_REPORT_AS",
    "EXTINCT", "EXTINCT_YEAR",
    "BIO_CONCEPT_CODE",
    "FOUR_LETTER_CODE",
    "PRI_COM_NAME_INDXD",
    "AUTHORITY_CODE", "AUTHORITY_VER",
}


def _classify_column(header: str):
    """
    Return (kind, language_code) for a taxonomy column, or (None, None).

    kind is currently always "common". Group-name parsing from the separate
    `species_groups` sheet is a follow-up — the sheet uses ISO codes whereas
    `full_sparse` uses display names, so aligning them needs a mapping pass.
    """
    if not header:
        return None, None
    raw = header.strip()
    if not raw:
        return None, None
    if raw.upper() in _CORE_COLUMNS:
        return None, None
    return "common", raw


class DetectionUploadResponse(BaseModel):
    """Response model for detection CSV upload."""
    success: bool
    detections_added: int
    detections_skipped: int
    species_created: int
    stations_matched: int
    message: str


AUTO_UPDATE_KEY = "auto_update_on_start"


@router.get("/auto-update-on-start")
async def get_auto_update_on_start(db: Session = Depends(get_db_dependency)):
    """
    Get the auto-update-on-start preference.

    Public (no auth) — Layout.tsx reads this on app load before login.
    Defaults to True for desktop mode, False for web mode, if not yet set.
    """
    setting = db.query(Setting).filter(Setting.key == AUTO_UPDATE_KEY).first()
    if setting:
        return {"enabled": setting.value.lower() in ("true", "1", "yes")}
    # Not set yet — return mode-based default
    default = app_settings.BWV_MODE == "desktop"
    return {"enabled": default}


@router.put("/auto-update-on-start")
async def set_auto_update_on_start(
    body: SettingUpdate,
    db: Session = Depends(get_db_dependency),
    current_user: dict = Depends(get_current_user),
):
    """
    Set the auto-update-on-start preference. Requires authentication.
    Send {"value": "true"} or {"value": "false"}.
    """
    enabled = body.value.lower() in ("true", "1", "yes")
    setting = db.query(Setting).filter(Setting.key == AUTO_UPDATE_KEY).first()
    if setting:
        setting.value = str(enabled).lower()
    else:
        setting = Setting(
            key=AUTO_UPDATE_KEY,
            value=str(enabled).lower(),
            data_type="bool",
            description="Auto-sync stations when the app starts",
        )
        db.add(setting)
    db.commit()
    return {"enabled": enabled}


@router.get("/", response_model=List[SettingResponse])
async def get_all_settings(
    db: Session = Depends(get_db_dependency),
    current_user: dict = Depends(get_current_user)
):
    """
    Get all application settings. Requires authentication.
    Filters out sensitive keys like password hashes.
    """
    SENSITIVE_KEYS = {"config_password_hash"}
    settings = db.query(Setting).all()
    return [SettingResponse(
        key=s.key,
        value=s.value,
        data_type=s.data_type,
        description=s.description
    ) for s in settings if s.key not in SENSITIVE_KEYS]


@router.get("/{key}", response_model=SettingResponse)
async def get_setting(
    key: str,
    db: Session = Depends(get_db_dependency),
    current_user: dict = Depends(get_current_user)
):
    """
    Get a specific setting by key.
    """
    setting = db.query(Setting).filter(Setting.key == key).first()
    if not setting:
        raise HTTPException(status_code=404, detail=f"Setting '{key}' not found")

    return SettingResponse(
        key=setting.key,
        value=setting.value,
        data_type=setting.data_type,
        description=setting.description
    )


@router.put("/{key}", response_model=SettingResponse)
async def update_setting(
    key: str,
    setting_data: SettingUpdate,
    db: Session = Depends(get_db_dependency),
    current_user: dict = Depends(get_current_user)
):
    """
    Update or create a setting.
    """
    setting = db.query(Setting).filter(Setting.key == key).first()

    if setting:
        setting.value = setting_data.value
        if setting_data.data_type:
            setting.data_type = setting_data.data_type
        if setting_data.description:
            setting.description = setting_data.description
    else:
        setting = Setting(
            key=key,
            value=setting_data.value,
            data_type=setting_data.data_type or "str",
            description=setting_data.description
        )
        db.add(setting)

    db.commit()
    db.refresh(setting)

    # Keep the taxonomy translation cache's app-wide language in sync when
    # the generic PUT path is used.
    if key == _tx.TAXONOMY_LANGUAGE_KEY:
        _tx.set_app_language(setting.value)

    return SettingResponse(
        key=setting.key,
        value=setting.value,
        data_type=setting.data_type,
        description=setting.description
    )


@router.delete("/{key}")
async def delete_setting(
    key: str,
    db: Session = Depends(get_db_dependency),
    current_user: dict = Depends(get_current_user)
):
    """
    Delete a setting.
    """
    setting = db.query(Setting).filter(Setting.key == key).first()
    if not setting:
        raise HTTPException(status_code=404, detail=f"Setting '{key}' not found")

    db.delete(setting)
    db.commit()

    return {"success": True, "message": f"Setting '{key}' deleted"}


def _iter_taxonomy_rows(file_content: bytes, filename: str):
    """
    Yield (headers, row_dict) tuples from an uploaded taxonomy file.

    Supports eBird's CSV export and the multilingual XLSX export. The XLSX
    path uses openpyxl read_only mode so memory stays bounded on large files.
    """
    lower = (filename or "").lower()
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise HTTPException(
                status_code=500,
                detail="openpyxl is not installed on the server; Excel (.xlsx) upload is unavailable",
            )
        wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        try:
            ws = wb.active
            rows = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows)
            except StopIteration:
                raise HTTPException(status_code=400, detail="Spreadsheet has no data")
            headers = [(str(h).strip() if h is not None else "") for h in header_row]
            yield headers, None  # first yield = headers
            for raw in rows:
                values = list(raw) + [None] * max(0, len(headers) - len(raw))
                record = {}
                for i, h in enumerate(headers):
                    if not h:
                        continue
                    v = values[i]
                    if v is None:
                        record[h] = ""
                    else:
                        record[h] = str(v).strip()
                yield headers, record
        finally:
            wb.close()
        return

    # CSV fallback
    text = None
    for encoding in ("utf-8", "latin-1", "cp1252"):
        try:
            text = file_content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise HTTPException(status_code=400, detail="Unable to decode CSV file")
    reader = csv.DictReader(io.StringIO(text))
    headers = reader.fieldnames or []
    if not headers:
        raise HTTPException(status_code=400, detail="CSV has no headers")
    yield headers, None
    for record in reader:
        clean = {k: (v.strip() if isinstance(v, str) else v) for k, v in record.items()}
        yield headers, clean


@router.post("/ebird-taxonomy", response_model=TaxonomyUploadResponse)
async def upload_ebird_taxonomy(
    file: UploadFile = File(...),
    db: Session = Depends(get_db_dependency),
    current_user: dict = Depends(get_current_user)
):
    """
    Upload eBird taxonomy CSV or XLSX to populate species codes and localized
    common names.

    Accepted files:
    - eBird taxonomy CSV (English only, fills SPECIES_CODE / family / order)
    - eBird multilingual taxonomy XLSX (also populates per-language common
      names so users can view plots with common names in their language)

    Required columns:
    - SPECIES_CODE (or SPECIES_CD)
    - SCI_NAME (or SCIENTIFIC_NAME)

    Optional columns:
    - PRIMARY_COM_NAME (English common name)
    - ORDER1 / FAMILY (taxonomy)
    - SPECIES_GROUP (English group name)
    - Any per-language columns (e.g. `es`, `de`, `ja_JP`, or `es_COM_NAME`,
      `es_GROUP_NAME`) — detected automatically.
    """
    filename = file.filename or ""
    if not (filename.lower().endswith(".csv") or filename.lower().endswith(".xlsx") or filename.lower().endswith(".xlsm")):
        raise HTTPException(status_code=400, detail="File must be a CSV or XLSX")

    try:
        content = await file.read()
        row_iter = _iter_taxonomy_rows(content, filename)
        headers, _ = next(row_iter)

        # Map the fixed columns we care about.
        col_map = {}
        common_name_cols_by_lang: Dict[str, str] = {}
        group_name_cols_by_lang: Dict[str, str] = {}

        for h in headers:
            u = h.upper().strip()
            if u in ("SPECIES_CODE", "SPECIES_CD", "SPECIESCODE"):
                col_map["species_code"] = h
            elif u in ("SCI_NAME", "SCIENTIFIC_NAME", "SCINAME"):
                col_map["scientific_name"] = h
            elif u in ("PRIMARY_COM_NAME", "COMMON_NAME", "COMNAME"):
                col_map["common_name"] = h
            elif u in ("ORDER1", "ORDER_NAME", "ORDER"):
                col_map["order"] = h
            elif u in ("FAMILY", "FAMILY_NAME"):
                col_map["family"] = h
            elif u in ("SPECIES_GROUP",):
                col_map["species_group"] = h
            else:
                kind, lang = _classify_column(h)
                if kind == "common" and lang:
                    common_name_cols_by_lang[lang] = h
                elif kind == "group" and lang:
                    group_name_cols_by_lang[lang] = h

        if "species_code" not in col_map or "scientific_name" not in col_map:
            raise HTTPException(
                status_code=400,
                detail="File must contain SPECIES_CODE and SCI_NAME columns",
            )

        updated_count = 0
        created_count = 0
        families_added = 0
        total_rows = 0
        families_seen = set()
        translations_imported = 0
        languages_seen = set()

        # If the file contains language columns, we replace the entire translation
        # table since the data is inherently keyed by this file's contents.
        if common_name_cols_by_lang or group_name_cols_by_lang:
            db.query(TaxonomyTranslation).delete()
            db.commit()

        for row in row_iter:
            _, record = row
            if record is None:
                continue

            species_code = (record.get(col_map["species_code"]) or "").strip()
            scientific_name = (record.get(col_map["scientific_name"]) or "").strip()
            if not species_code or not scientific_name:
                continue

            total_rows += 1

            common_name = (record.get(col_map.get("common_name", "")) or "").strip() if "common_name" in col_map else ""
            family = (record.get(col_map.get("family", "")) or "").strip() if "family" in col_map else ""
            order = (record.get(col_map.get("order", "")) or "").strip() if "order" in col_map else ""
            if family:
                families_seen.add(family)

            species = db.query(Species).filter(
                Species.scientific_name.ilike(scientific_name)
            ).first()

            if species:
                species.ebird_code = species_code
                if family and not species.family:
                    families_added += 1
                if family:
                    species.family = family
                if order:
                    species.order = order
                if common_name and not species.common_name:
                    species.common_name = common_name
                updated_count += 1
            else:
                species = Species(
                    scientific_name=scientific_name,
                    common_name=common_name or scientific_name,
                    species_id=None,
                    ebird_code=species_code,
                    family=family or None,
                    order=order or None,
                )
                db.add(species)
                db.flush()  # get species.id for translation rows below
                created_count += 1
                if family:
                    families_added += 1

            # Persist per-language common and group names keyed on species.id
            for lang, col in common_name_cols_by_lang.items():
                value = (record.get(col) or "").strip()
                if not value:
                    continue
                group_col = group_name_cols_by_lang.get(lang)
                group_value = (record.get(group_col) or "").strip() if group_col else None
                db.add(TaxonomyTranslation(
                    species_id=species.id,
                    language_code=lang,
                    common_name=value,
                    group_name=group_value or None,
                ))
                translations_imported += 1
                languages_seen.add(lang)

            # Group-only languages (rare — if a file has group column but no common name column)
            for lang, col in group_name_cols_by_lang.items():
                if lang in common_name_cols_by_lang:
                    continue  # already handled above
                group_value = (record.get(col) or "").strip()
                if not group_value:
                    continue
                db.add(TaxonomyTranslation(
                    species_id=species.id,
                    language_code=lang,
                    common_name=None,
                    group_name=group_value,
                ))
                translations_imported += 1
                languages_seen.add(lang)

            if (updated_count + created_count) % 1000 == 0:
                db.commit()

        db.commit()

        # Refresh in-memory caches so the API starts returning localized
        # names immediately without a server restart.
        _tx.load_cache(db)

        # Update taxonomy stats setting
        stats_setting = db.query(Setting).filter(Setting.key == "ebird_taxonomy_stats").first()
        import json
        stats = {
            "last_updated": datetime.utcnow().isoformat(),
            "species_with_codes": db.query(Species).filter(Species.ebird_code.isnot(None)).count(),
            "species_with_families": db.query(Species).filter(Species.family.isnot(None)).count(),
            "total_species": db.query(Species).count(),
            "unique_families": len(families_seen),
            "ebird_entries": total_rows,
            "languages": sorted(languages_seen),
            "translations_imported": translations_imported,
        }
        if stats_setting:
            stats_setting.value = json.dumps(stats)
            stats_setting.data_type = "json"
        else:
            db.add(Setting(
                key="ebird_taxonomy_stats",
                value=json.dumps(stats),
                data_type="json",
                description="eBird taxonomy import statistics",
            ))
        db.commit()

        languages_list = sorted(languages_seen)
        msg = (
            f"Processed {total_rows} eBird species. Updated {updated_count} existing, "
            f"created {created_count} new. {families_added} species now have family data. "
            f"Found {len(families_seen)} unique families."
        )
        if translations_imported:
            msg += f" Imported {translations_imported} localized names across {len(languages_list)} language(s)."

        return TaxonomyUploadResponse(
            success=True,
            species_updated=updated_count,
            species_created=created_count,
            families_added=families_added,
            total_ebird_species=total_rows,
            languages_imported=languages_list,
            translations_imported=translations_imported,
            message=msg,
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Error processing taxonomy file")
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")


@router.get("/ebird-taxonomy/stats")
async def get_taxonomy_stats(db: Session = Depends(get_db_dependency)):
    """
    Get eBird taxonomy statistics, including the list of languages available
    for common-name localization and the currently selected language.
    """
    import json

    stats_setting = db.query(Setting).filter(Setting.key == 'ebird_taxonomy_stats').first()

    species_with_codes = db.query(Species).filter(Species.ebird_code.isnot(None)).count()
    species_with_families = db.query(Species).filter(Species.family.isnot(None)).count()
    total_species = db.query(Species).count()

    # Get stored stats for additional fields
    stored_stats = {}
    if stats_setting:
        try:
            stored_stats = json.loads(stats_setting.value)
        except (json.JSONDecodeError, TypeError):
            pass

    # Ensure translation cache reflects current DB state for languages list.
    _tx.ensure_loaded(db)

    languages = _tx.available_languages()
    current_language = _tx.current_language() or "en"

    return {
        'species_with_codes': species_with_codes,
        'species_with_families': species_with_families,
        'total_species': total_species,
        'coverage_percent': round(species_with_codes / total_species * 100, 1) if total_species > 0 else 0,
        'unique_families': stored_stats.get('unique_families', 0),
        'ebird_entries': stored_stats.get('ebird_entries', 0),
        'last_updated': stored_stats.get('last_updated'),
        'available_languages': languages,
        'current_language': current_language,
        'translations_imported': stored_stats.get('translations_imported', 0),
    }


@router.get("/ebird-taxonomy/language")
async def get_taxonomy_language(db: Session = Depends(get_db_dependency)):
    """
    Return the currently selected common-name language. Public (no auth) —
    layout and per-page requests need it to decide rendering.
    """
    _tx.ensure_loaded(db)
    return {
        "language": _tx.current_language() or "en",
        "available_languages": _tx.available_languages(),
    }


@router.put("/ebird-taxonomy/language")
async def set_taxonomy_language(
    body: SettingUpdate,
    db: Session = Depends(get_db_dependency),
    current_user: dict = Depends(get_current_user),
):
    """
    Set the current common-name language. Pass a language code present in the
    uploaded taxonomy (e.g. "es", "de", "ja_JP") or "en" to revert to the
    canonical English names.
    """
    code = (body.value or "").strip()
    if code and code != "en":
        _tx.ensure_loaded(db)
        if code not in _tx.available_languages():
            raise HTTPException(
                status_code=400,
                detail=f"Language '{code}' is not available. Upload a taxonomy file that includes it.",
            )

    setting = db.query(Setting).filter(Setting.key == _tx.TAXONOMY_LANGUAGE_KEY).first()
    if setting:
        setting.value = code
    else:
        setting = Setting(
            key=_tx.TAXONOMY_LANGUAGE_KEY,
            value=code,
            data_type="str",
            description="Language code for localized bird common names in plots",
        )
        db.add(setting)
    db.commit()
    _tx.set_app_language(code)
    return {"language": code or "en"}


@router.post("/detections/upload", response_model=DetectionUploadResponse)
async def upload_detections_csv(
    file: UploadFile = File(...),
    db: Session = Depends(get_db_dependency),
    current_user: dict = Depends(get_current_user)
):
    """
    Upload detection data from CSV file.

    Expected CSV columns:
    - Timestamp: Detection timestamp (e.g., "2025-12-22 19:49:42 -0500")
    - Common Name: Bird common name
    - Scientific Name: Bird scientific name
    - Latitude: Detection latitude
    - Longitude: Detection longitude
    - Station: Station name to match against existing stations
    - Confidence: Detection confidence (0-1)
    - Score: Detection score (optional)
    - Probability: Detection probability (optional)
    - Soundscape: URL to soundscape audio (optional)
    """
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="File must be a CSV")

    try:
        content = await file.read()
        # Try different encodings
        for encoding in ['utf-8', 'latin-1', 'cp1252']:
            try:
                text_content = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        else:
            raise HTTPException(status_code=400, detail="Unable to decode CSV file")

        # Parse CSV
        reader = csv.DictReader(io.StringIO(text_content))

        if not reader.fieldnames:
            raise HTTPException(status_code=400, detail="CSV has no headers")

        # Normalize column names
        col_map = {}
        for field in reader.fieldnames:
            field_lower = field.lower().strip()
            if 'timestamp' in field_lower:
                col_map['timestamp'] = field
            elif field_lower == 'common name':
                col_map['common_name'] = field
            elif field_lower == 'scientific name':
                col_map['scientific_name'] = field
            elif field_lower == 'latitude':
                col_map['latitude'] = field
            elif field_lower == 'longitude':
                col_map['longitude'] = field
            elif field_lower == 'station':
                col_map['station'] = field
            elif field_lower == 'confidence':
                col_map['confidence'] = field
            elif field_lower == 'score':
                col_map['score'] = field
            elif field_lower == 'soundscape':
                col_map['soundscape'] = field

        required = ['timestamp', 'scientific_name', 'station']
        missing = [r for r in required if r not in col_map]
        if missing:
            raise HTTPException(
                status_code=400,
                detail=f"CSV missing required columns: {', '.join(missing)}"
            )

        # Cache for stations and species
        station_cache = {}
        species_cache = {}

        detections_added = 0
        detections_skipped = 0
        species_created = 0
        stations_matched = set()

        for row in reader:
            try:
                # Parse timestamp
                timestamp_str = row.get(col_map['timestamp'], '').strip()
                if not timestamp_str:
                    detections_skipped += 1
                    continue

                # Parse various timestamp formats
                timestamp = None
                for fmt in [
                    '%Y-%m-%d %H:%M:%S %z',  # 2025-12-22 19:49:42 -0500
                    '%Y-%m-%d %H:%M:%S',      # 2025-12-22 19:49:42
                    '%Y-%m-%dT%H:%M:%S%z',    # ISO format
                    '%Y-%m-%dT%H:%M:%S',      # ISO without tz
                ]:
                    try:
                        timestamp = datetime.strptime(timestamp_str, fmt)
                        break
                    except ValueError:
                        continue

                if not timestamp:
                    logger.warning(f"Could not parse timestamp: {timestamp_str}")
                    detections_skipped += 1
                    continue

                # Get station
                station_name = row.get(col_map['station'], '').strip()
                if not station_name:
                    detections_skipped += 1
                    continue

                if station_name not in station_cache:
                    # Match station by name (partial match)
                    station = db.query(Station).filter(
                        Station.name.ilike(f"%{station_name}%")
                    ).first()
                    station_cache[station_name] = station

                station = station_cache[station_name]
                if not station:
                    detections_skipped += 1
                    continue

                stations_matched.add(station.id)

                # Get or create species
                scientific_name = row.get(col_map['scientific_name'], '').strip()
                common_name = row.get(col_map.get('common_name', ''), '').strip() if 'common_name' in col_map else ''

                if not scientific_name:
                    detections_skipped += 1
                    continue

                if scientific_name not in species_cache:
                    species = db.query(Species).filter(
                        Species.scientific_name.ilike(scientific_name)
                    ).first()

                    if not species:
                        # Create new species
                        species = Species(
                            scientific_name=scientific_name,
                            common_name=common_name or scientific_name,
                            species_id=None  # Manual import, no BirdWeather ID
                        )
                        db.add(species)
                        db.flush()
                        species_created += 1

                    species_cache[scientific_name] = species

                species = species_cache[scientific_name]

                # Parse coordinates
                try:
                    latitude = float(row.get(col_map.get('latitude', ''), 0) or 0)
                    longitude = float(row.get(col_map.get('longitude', ''), 0) or 0)
                except (ValueError, TypeError):
                    latitude = station.latitude or 0
                    longitude = station.longitude or 0

                # Parse confidence
                try:
                    confidence = float(row.get(col_map.get('confidence', ''), 0) or 0)
                except (ValueError, TypeError):
                    confidence = 0.0

                # Check for duplicate detection (same station, species, timestamp)
                existing = db.query(Detection).filter(
                    Detection.station_id == station.id,
                    Detection.species_id == species.id,
                    Detection.timestamp == timestamp
                ).first()

                if existing:
                    detections_skipped += 1
                    continue

                # Create detection
                detection = Detection(
                    station_id=station.id,
                    species_id=species.id,
                    detection_id=None,  # Manual upload, no BirdWeather ID
                    timestamp=timestamp,
                    confidence=confidence,
                    latitude=latitude,
                    longitude=longitude,
                    detection_date=timestamp.date(),
                    detection_hour=timestamp.hour
                )
                db.add(detection)
                detections_added += 1

                # Commit periodically to avoid memory issues
                if detections_added % 500 == 0:
                    db.commit()

            except Exception as e:
                logger.warning(f"Error processing row: {e}")
                detections_skipped += 1
                continue

        db.commit()

        return DetectionUploadResponse(
            success=True,
            detections_added=detections_added,
            detections_skipped=detections_skipped,
            species_created=species_created,
            stations_matched=len(stations_matched),
            message=f"Added {detections_added} detections. Skipped {detections_skipped}. Created {species_created} new species. Matched {len(stations_matched)} stations."
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error uploading detections: {e}")
        raise HTTPException(status_code=500, detail=f"Error processing CSV: {str(e)}")


@router.post("/detections/upload-stream")
async def upload_detections_csv_stream(
    file: UploadFile = File(...),
    db: Session = Depends(get_db_dependency),
    current_user: dict = Depends(get_current_user)
):
    """
    Upload detection data from CSV with streaming progress via NDJSON.

    Emits line-by-line progress events so the frontend can show a progress bar.
    """
    from fastapi.responses import StreamingResponse
    import json

    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="File must be a CSV")

    content = await file.read()

    # Try different encodings
    text_content = None
    for encoding in ['utf-8', 'latin-1', 'cp1252']:
        try:
            text_content = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue

    if text_content is None:
        raise HTTPException(status_code=400, detail="Unable to decode CSV file")

    def generate_progress():
        # Count total lines (excluding header)
        lines = text_content.strip().split('\n')
        total_lines = max(len(lines) - 1, 0)  # subtract header

        yield json.dumps({
            "type": "start",
            "total_lines": total_lines
        }) + "\n"

        reader = csv.DictReader(io.StringIO(text_content))

        if not reader.fieldnames:
            yield json.dumps({
                "type": "error",
                "message": "CSV has no headers"
            }) + "\n"
            return

        # Normalize column names
        col_map = {}
        for field in reader.fieldnames:
            field_lower = field.lower().strip()
            if 'timestamp' in field_lower:
                col_map['timestamp'] = field
            elif field_lower == 'common name':
                col_map['common_name'] = field
            elif field_lower == 'scientific name':
                col_map['scientific_name'] = field
            elif field_lower == 'latitude':
                col_map['latitude'] = field
            elif field_lower == 'longitude':
                col_map['longitude'] = field
            elif field_lower == 'station':
                col_map['station'] = field
            elif field_lower == 'confidence':
                col_map['confidence'] = field
            elif field_lower == 'score':
                col_map['score'] = field
            elif field_lower == 'soundscape':
                col_map['soundscape'] = field

        required = ['timestamp', 'scientific_name', 'station']
        missing = [r for r in required if r not in col_map]
        if missing:
            yield json.dumps({
                "type": "error",
                "message": f"CSV missing required columns: {', '.join(missing)}"
            }) + "\n"
            return

        station_cache = {}
        species_cache = {}
        detections_added = 0
        detections_skipped = 0
        species_created = 0
        stations_matched = set()
        line_number = 0

        for row in reader:
            line_number += 1
            try:
                timestamp_str = row.get(col_map['timestamp'], '').strip()
                if not timestamp_str:
                    detections_skipped += 1
                    continue

                timestamp = None
                for fmt in [
                    '%Y-%m-%d %H:%M:%S %z',
                    '%Y-%m-%d %H:%M:%S',
                    '%Y-%m-%dT%H:%M:%S%z',
                    '%Y-%m-%dT%H:%M:%S',
                ]:
                    try:
                        timestamp = datetime.strptime(timestamp_str, fmt)
                        break
                    except ValueError:
                        continue

                if not timestamp:
                    detections_skipped += 1
                    continue

                station_name = row.get(col_map['station'], '').strip()
                if not station_name:
                    detections_skipped += 1
                    continue

                if station_name not in station_cache:
                    station = db.query(Station).filter(
                        Station.name.ilike(f"%{station_name}%")
                    ).first()
                    station_cache[station_name] = station

                station = station_cache[station_name]
                if not station:
                    detections_skipped += 1
                    continue

                stations_matched.add(station.id)

                scientific_name = row.get(col_map['scientific_name'], '').strip()
                common_name = row.get(col_map.get('common_name', ''), '').strip() if 'common_name' in col_map else ''

                if not scientific_name:
                    detections_skipped += 1
                    continue

                if scientific_name not in species_cache:
                    species = db.query(Species).filter(
                        Species.scientific_name.ilike(scientific_name)
                    ).first()

                    if not species:
                        species = Species(
                            scientific_name=scientific_name,
                            common_name=common_name or scientific_name,
                            species_id=None
                        )
                        db.add(species)
                        db.flush()
                        species_created += 1

                    species_cache[scientific_name] = species

                species = species_cache[scientific_name]

                try:
                    latitude = float(row.get(col_map.get('latitude', ''), 0) or 0)
                    longitude = float(row.get(col_map.get('longitude', ''), 0) or 0)
                except (ValueError, TypeError):
                    latitude = station.latitude or 0
                    longitude = station.longitude or 0

                try:
                    confidence = float(row.get(col_map.get('confidence', ''), 0) or 0)
                except (ValueError, TypeError):
                    confidence = 0.0

                existing = db.query(Detection).filter(
                    Detection.station_id == station.id,
                    Detection.species_id == species.id,
                    Detection.timestamp == timestamp
                ).first()

                if existing:
                    detections_skipped += 1
                    continue

                detection = Detection(
                    station_id=station.id,
                    species_id=species.id,
                    detection_id=None,
                    timestamp=timestamp,
                    confidence=confidence,
                    latitude=latitude,
                    longitude=longitude,
                    detection_date=timestamp.date(),
                    detection_hour=timestamp.hour
                )
                db.add(detection)
                detections_added += 1

                if detections_added % 500 == 0:
                    db.commit()

            except Exception as e:
                logger.warning(f"Error processing row {line_number}: {e}")
                detections_skipped += 1
                continue

            # Emit progress every 50 lines to avoid overwhelming the stream
            if line_number % 50 == 0:
                yield json.dumps({
                    "type": "progress",
                    "lines_processed": line_number,
                    "total_lines": total_lines,
                    "detections_added": detections_added,
                    "detections_skipped": detections_skipped
                }) + "\n"

        db.commit()

        yield json.dumps({
            "type": "complete",
            "success": True,
            "detections_added": detections_added,
            "detections_skipped": detections_skipped,
            "species_created": species_created,
            "stations_matched": len(stations_matched),
            "lines_processed": line_number,
            "total_lines": total_lines,
            "message": f"Added {detections_added} detections. Skipped {detections_skipped}. Created {species_created} new species. Matched {len(stations_matched)} stations."
        }) + "\n"

    return StreamingResponse(
        generate_progress(),
        media_type="application/x-ndjson"
    )


# ============================================================================
# Database export / import
# ============================================================================

def _sqlite_path_from_url(url: str) -> Optional[str]:
    """Return the on-disk path of a SQLite DATABASE_URL, or None if not SQLite."""
    if not url or not url.startswith("sqlite"):
        return None
    # Forms: "sqlite:///relative/path.db", "sqlite:////abs/path.db"
    path = url.split("sqlite:///", 1)[-1]
    if not path:
        return None
    return os.path.abspath(path)


@router.get("/db/info")
async def get_db_info(current_user: dict = Depends(get_current_user)):
    """
    Report whether the database supports file-level export/import (SQLite only).
    """
    path = _sqlite_path_from_url(app_settings.DATABASE_URL)
    if not path:
        return {"supported": False, "engine": app_settings.DATABASE_URL.split("://")[0]}
    exists = os.path.exists(path)
    size_bytes = os.path.getsize(path) if exists else 0
    return {
        "supported": True,
        "engine": "sqlite",
        "path": path,
        "exists": exists,
        "size_bytes": size_bytes,
    }


@router.get("/db/export")
async def export_database(
    db: Session = Depends(get_db_dependency),
    current_user: dict = Depends(get_current_user),
):
    """
    Download a full snapshot of the SQLite database.

    The WAL is checkpointed into the main file and then a consistent copy is
    streamed as an attachment. Works for SQLite only — returns 400 otherwise.
    """
    path = _sqlite_path_from_url(app_settings.DATABASE_URL)
    if not path:
        raise HTTPException(
            status_code=400,
            detail="Database export is only supported for SQLite deployments.",
        )
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Database file not found.")

    # Fold WAL into the main .db so the file we copy is consistent and
    # self-contained. TRUNCATE also reclaims the WAL file.
    from sqlalchemy import text as _sa_text
    try:
        db.execute(_sa_text("PRAGMA wal_checkpoint(TRUNCATE)"))
        db.commit()
    except Exception as e:
        logger.warning("wal_checkpoint failed (continuing): %s", e)

    # Copy to a temp file so streaming is not racing with concurrent writers.
    suffix = ".sqlite"
    tmp_fd, tmp_path = tempfile.mkstemp(prefix="bwv-export-", suffix=suffix)
    os.close(tmp_fd)
    try:
        shutil.copyfile(path, tmp_path)
    except Exception as e:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        raise HTTPException(status_code=500, detail=f"Failed to snapshot DB: {e}")

    timestamp = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    download_name = f"birdweatherviz-backup-{timestamp}.sqlite"

    # FileResponse streams and closes the handle. We rely on the OS to let us
    # delete later — we schedule cleanup via a background task.
    from fastapi import BackgroundTasks
    background = BackgroundTasks()
    background.add_task(lambda p=tmp_path: os.path.exists(p) and os.remove(p))
    return FileResponse(
        tmp_path,
        media_type="application/x-sqlite3",
        filename=download_name,
        background=background,
    )


class DbImportResponse(BaseModel):
    success: bool
    message: str
    restored_bytes: int
    backup_path: Optional[str] = None


@router.post("/db/import", response_model=DbImportResponse)
async def import_database(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
):
    """
    Replace the current SQLite database with an uploaded backup.

    Safety:
    - The uploaded file must be a SQLite database and must contain our core
      tables (`species`, `station`, `detection`). Anything else is rejected.
    - Before overwriting, the existing DB is renamed to `<name>.pre-restore-<ts>`
      so a rollback is possible from the filesystem.
    - After swap, the SQLAlchemy engine pool is disposed so new sessions open
      the restored file. The app should be restarted for a fully clean state,
      but most endpoints will pick up the new DB immediately.
    """
    path = _sqlite_path_from_url(app_settings.DATABASE_URL)
    if not path:
        raise HTTPException(
            status_code=400,
            detail="Database import is only supported for SQLite deployments.",
        )

    fn = (file.filename or "").lower()
    if not (fn.endswith(".sqlite") or fn.endswith(".db") or fn.endswith(".sqlite3")):
        raise HTTPException(
            status_code=400,
            detail="File must be a SQLite database (.sqlite, .sqlite3, or .db).",
        )

    content = await file.read()
    if len(content) < 100 or not content.startswith(b"SQLite format 3\x00"):
        raise HTTPException(
            status_code=400,
            detail="Uploaded file is not a valid SQLite database.",
        )

    # Validate schema — open read-only, check our tables exist.
    tmp_fd, tmp_path = tempfile.mkstemp(prefix="bwv-import-", suffix=".sqlite")
    try:
        with os.fdopen(tmp_fd, "wb") as f:
            f.write(content)

        import sqlite3
        required = {"species", "station", "detection", "setting"}
        try:
            conn = sqlite3.connect(f"file:{tmp_path}?mode=ro", uri=True)
        except sqlite3.Error as e:
            raise HTTPException(status_code=400, detail=f"Could not open uploaded DB: {e}")
        try:
            names = {
                row[0]
                for row in conn.execute(
                    "SELECT name FROM sqlite_master WHERE type='table'"
                ).fetchall()
            }
        finally:
            conn.close()
        missing = required - names
        if missing:
            raise HTTPException(
                status_code=400,
                detail=f"Uploaded DB is missing required tables: {', '.join(sorted(missing))}",
            )

        # Dispose the current engine pool and sidecars (WAL, SHM) before swap.
        from app.db.session import engine
        try:
            engine.dispose()
        except Exception as e:
            logger.warning("Engine dispose failed (continuing): %s", e)

        target = Path(path)
        target.parent.mkdir(parents=True, exist_ok=True)

        # Move current DB + sidecars out of the way.
        backup_stamp = time.strftime("%Y%m%d-%H%M%S")
        backup_path = None
        if target.exists():
            backup_path = f"{target}.pre-restore-{backup_stamp}"
            os.rename(target, backup_path)
            for sidecar in (f"{target}-wal", f"{target}-shm"):
                if os.path.exists(sidecar):
                    try:
                        os.rename(sidecar, f"{sidecar}.pre-restore-{backup_stamp}")
                    except OSError:
                        pass

        shutil.move(tmp_path, target)

        # Reload the in-memory taxonomy cache + language from the new DB.
        from app.db.session import SessionLocal

        db = SessionLocal()
        try:
            _tx.load_cache(db)
            _tx.load_app_language(db)
        finally:
            db.close()

        return DbImportResponse(
            success=True,
            message="Database restored. Reload the page to see the restored data.",
            restored_bytes=len(content),
            backup_path=backup_path,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Error importing database")
        raise HTTPException(status_code=500, detail=f"Import failed: {e}")
    finally:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except OSError:
                pass
